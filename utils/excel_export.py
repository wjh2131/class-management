from models import User, Class, Achievement, Activity, Leave, ClassFund, db, query_get
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from datetime import datetime
import os




def export_class_contact_list(class_id=None, filename=None):
    """
    导出班级通讯录为Excel文件

    Args:
        class_id: 班级ID，如果为None则导出所有用户
        filename: 自定义文件名，如果为None则自动生成

    Returns:
        str: 生成的Excel文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "班级通讯录"

    # 定义表头
    headers = ['学号', '姓名', '角色', '班级', '手机号', '邮箱', '注册时间', '状态']
    ws.append(headers)

    # 设置表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 查询数据
    if class_id:
        users = User.query.filter_by(class_id=class_id).order_by(User.student_id).all()
        class_obj = query_get(Class, class_id)
        class_name = class_obj.name if class_obj else '未知班级'
    else:
        users = User.query.order_by(User.created_at.desc()).all()
        class_name = '全部'

    # 状态映射
    status_map = {
        'pending': '待审核',
        'approved': '已通过',
        'rejected': '已拒绝'
    }

    # 填充数据
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    for idx, user in enumerate(users, start=2):
        user_class = user.student_class
        class_name_str = user_class.name if user_class else '未分配'

        row_data = [
            user.student_id,
            user.real_name,
            user.role_name,
            class_name_str,
            user.phone,
            user.email,
            user.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            status_map.get(user.status, user.status)
        ]
        ws.append(row_data)

        # 设置数据行样式
        for cell in ws[idx]:
            cell.alignment = data_alignment if cell.column != 3 else center_alignment
            cell.border = thin_border

            #  alternating row colors
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 自动调整列宽
    column_widths = {
        'A': 15,  # 学号
        'B': 12,  # 姓名
        'C': 10,  # 角色
        'D': 20,  # 班级
        'E': 15,  # 手机号
        'F': 25,  # 邮箱
        'G': 20,  # 注册时间
        'H': 10   # 状态
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加统计信息
    stats_row = len(users) + 3
    ws.merge_cells(f'A{stats_row}:H{stats_row}')
    stats_cell = ws[f'A{stats_row}']
    stats_cell.value = f"总计: {len(users)} 人 | 导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    stats_cell.font = Font(bold=True, italic=True, color='666666')
    stats_cell.alignment = Alignment(horizontal='right')

    # 生成文件名
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if class_id:
            safe_class_name = class_name.replace('/', '_').replace('\\', '_')
            filename = f'class_contact_{safe_class_name}_{timestamp}.xlsx'
        else:
            filename = f'all_users_contact_{timestamp}.xlsx'

    # 确保目录存在
    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports')
    os.makedirs(export_dir, exist_ok=True)

    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return filepath


def export_fund_report(class_id, year, month, filename=None):
    """
    导出班费月度报表

    Args:
        class_id: 班级ID
        year: 年份
        month: 月份
        filename: 自定义文件名

    Returns:
        str: 生成的Excel文件路径
    """
    from models import ClassFund

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月班费报表"

    # 标题
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    class_obj = query_get(Class, class_id)
    class_name = class_obj.name if class_obj else '未知班级'
    title_cell.value = f"{class_name} {year}年{month}月班费收支明细表"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 表头
    headers = ['日期', '描述', '类型', '金额', '经办人', '记录时间']
    ws.append(headers)

    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 查询数据
    from datetime import date
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    funds = ClassFund.query.filter(
        ClassFund.class_id == class_id,
        ClassFund.date >= start_date,
        ClassFund.date < end_date
    ).order_by(ClassFund.date).all()

    # 填充数据
    total_income = 0
    total_expense = 0

    data_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')

    for idx, fund in enumerate(funds, start=3):
        amount = fund.amount
        if fund.type == 'income':
            total_income += amount
            display_amount = f"+{amount:.2f}"
            cell_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        else:
            total_expense += amount
            display_amount = f"-{amount:.2f}"
            cell_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        type_text = '收入' if fund.type == 'income' else '支出'

        row_data = [
            fund.date.strftime('%Y-%m-%d'),
            fund.description,
            type_text,
            display_amount,
            fund.operator,
            fund.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        ws.append(row_data)

        # 设置样式
        for cell in ws[idx]:
            cell.border = thin_border
            if cell.column == 4:  # 金额列
                cell.alignment = number_alignment
            else:
                cell.alignment = data_alignment
            cell.fill = cell_fill

    # 添加汇总行
    summary_row = len(funds) + 3
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    ws[f'A{summary_row}'] = '合计'
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'A{summary_row}'].alignment = Alignment(horizontal='right')
    ws[f'A{summary_row}'].border = thin_border

    ws[f'D{summary_row}'] = f"{total_income - total_expense:.2f}"
    ws[f'D{summary_row}'].font = Font(bold=True, color='FF0000' if (total_income - total_expense) < 0 else '008000')
    ws[f'D{summary_row}'].alignment = number_alignment
    ws[f'D{summary_row}'].border = thin_border

    ws.merge_cells(f'E{summary_row}:F{summary_row}')
    ws[f'E{summary_row}'] = f"收入: {total_income:.2f} | 支出: {total_expense:.2f} | 余额: {total_income - total_expense:.2f}"
    ws[f'E{summary_row}'].font = Font(bold=True, italic=True)
    ws[f'E{summary_row}'].alignment = Alignment(horizontal='right')
    ws[f'E{summary_row}'].border = thin_border

    # 设置列宽
    column_widths = {
        'A': 12,
        'B': 30,
        'C': 10,
        'D': 15,
        'E': 12,
        'F': 18
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 生成文件名
    if not filename:
        filename = f'fund_report_{class_name}_{year}_{month}.xlsx'

    # 保存文件
    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports')
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

def export_achievements_excel(class_id=None, status='approved'):
    """导出成果记录"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成果记录"

    headers = ['标题', '类别', '提交人', '提交时间', '状态', '审核意见']
    ws.append(headers)

    header_fill = PatternFill(start_color='667eea', end_color='764ba2', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    query = Achievement.query.filter_by(status=status)
    if class_id:
        query = query.join(User).filter(User.class_id == class_id)

    achievements = query.order_by(Achievement.submitted_at.desc()).all()

    for ach in achievements:
        ws.append([
            ach.title,
            ach.category_name,
            ach.submitter_name,
            ach.submitted_at.strftime('%Y-%m-%d %H:%M'),
            ach.status_name,
            ach.review_comment or ''
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'achievements_{timestamp}.xlsx'

    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports')
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return filepath


def export_leave_records(class_id=None, status='approved'):
    """导出请假记录"""
    wb = Workbook()
    ws = wb.active
    ws.title = "请假记录"

    headers = ['学生姓名', '请假类型', '开始日期', '结束日期', '天数', '原因', '状态']
    ws.append(headers)

    header_fill = PatternFill(start_color='667eea', end_color='764ba2', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    query = Leave.query.filter_by(status=status)
    if class_id:
        query = query.join(User, Leave.student_id == User.id).filter(User.class_id == class_id)


    leaves = query.order_by(Leave.applied_at.desc()).all()

    for leave in leaves:
        ws.append([
            leave.student.real_name,
            leave.leave_type,
            leave.start_date.strftime('%Y-%m-%d'),
            leave.end_date.strftime('%Y-%m-%d'),
            leave.days,
            leave.reason,
            leave.status_name
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'leave_records_{timestamp}.xlsx'

    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports')
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return filepath


def export_activity_list(activity_type=None):
    """导出活动列表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "活动列表"

    headers = ['活动名称', '类型', '地点', '开始时间', '结束时间', '报名人数', '状态']
    ws.append(headers)

    header_fill = PatternFill(start_color='667eea', end_color='764ba2', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    query = Activity.query
    if activity_type:
        query = query.filter_by(activity_type=activity_type)

    activities = query.order_by(Activity.start_time.desc()).all()

    for activity in activities:
        ws.append([
            activity.title,
            activity.activity_type_name,
            activity.location or '未定',
            activity.start_time.strftime('%Y-%m-%d %H:%M'),
            activity.end_time.strftime('%Y-%m-%d %H:%M'),
            activity.registered_count,
            activity.status_name
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'activities_{timestamp}.xlsx'

    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports')
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return filename



def export_activity_list(activity_type=None):
    """导出活动列表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "活动列表"

    headers = ['活动名称', '类型', '地点', '开始时间', '结束时间', '报名人数', '状态']
    ws.append(headers)

    header_fill = PatternFill(start_color='667eea', end_color='764ba2', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    query = Activity.query
    if activity_type:
        query = query.filter_by(activity_type=activity_type)

    activities = query.order_by(Activity.start_time.desc()).all()

    for activity in activities:
        ws.append([
            activity.title,
            activity.activity_type_name,
            activity.location or '未定',
            activity.start_time.strftime('%Y-%m-%d %H:%M'),
            activity.end_time.strftime('%Y-%m-%d %H:%M'),
            activity.registered_count,
            activity.status_name
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'activities_{timestamp}.xlsx'
    wb.save(filename)

    return filename



